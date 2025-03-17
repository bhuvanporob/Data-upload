import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.figure_factory as ff
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Border, Side
import datetime as dt

def apply_excel_formatting(writer, sheet_name, df):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    alternate_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    green_fill = PatternFill(start_color="B3E7B5", end_color="B3E7B5", fill_type="solid")
    red_fill = PatternFill(start_color="FDB1B1", end_color="FDB1B1", fill_type="solid")
    green_font = Font(color="006100")
    red_font = Font(color="9C0006")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        fill = white_fill if row_idx % 2 == 0 else alternate_fill
        for cell in row:
            cell.fill = fill
            cell.border = thin_border

    inv_col_idx = None
    for col_idx, cell in enumerate(worksheet[1], start=1):
        if cell.value in ["INV_per", "IND_per"]:
            inv_col_idx = col_idx
            break

    if inv_col_idx:
        for row_idx, value in enumerate(df["INV_per" if "INV_per" in df.columns else "IND_per"], start=2):
            cell = worksheet.cell(row=row_idx, column=inv_col_idx)
            if isinstance(value, (int, float)):
                if value > 7:
                    cell.fill = red_fill
                    cell.font = red_font
                else:
                    cell.fill = green_fill
                    cell.font = green_font
            cell.border = thin_border

def clean_data(df):
    df = df[df['Test_status'].notnull()]
    df_wun = df[~df['User_name'].isin(['Service'])]
    df_wun = df_wun[~df_wun['Lab_name'].isin(['QC'])]
    df_clean = df_wun.assign(Truelab_id=df_wun['Truelab_id'].apply(lambda x: x.partition('-')[0]))
    df_clean['Chip_serial_no'] = df_clean['Chip_serial_no'].str[:2]
    df_clean = df_clean.drop_duplicates(keep='first')
    df_clean['Ct1'] = pd.to_numeric(df_clean['Ct1'], errors='coerce').fillna(0)
    df_clean['Ct2'] = pd.to_numeric(df_clean['Ct2'], errors='coerce').fillna(0)
    df_clean['Ct3'] = pd.to_numeric(df_clean['Ct3'], errors='coerce').fillna(0)
    # df_clean['Chip_serial_no'] = df_clean['Chip_serial_no'].astype(str)
    df_clean = df_clean.dropna(subset=['Chip_serial_no'])  # Remove NaNs
    df_clean['Chip_serial_no'] = df_clean['Chip_serial_no'].astype(str).str.strip()
    df_clean = df_clean[df_clean['Chip_serial_no'].str.len() > 1]  # Keep only values with at least 2 characters
    df_clean = df_clean[~df_clean['Chip_serial_no'].str[0].str.isdigit()]
    df_clean = df_clean[df_clean['Chip_serial_no'].str[1].str.isdigit()]
    masterlist = pd.read_excel('Feb Materlist.xlsx',sheet_name='Main data')
    df_merged = pd.merge(df_clean,masterlist[['Truelab_id','Zone','Account Owner','State','Customer Type']],on='Truelab_id',how='left')  
    return df_merged

def generate_excel(dataframes):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in dataframes.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_excel_formatting(writer, sheet_name, df)
    output.seek(0)
    return output

def main():
    st.title("CSV File Uploader and Viewer")
    
    uploaded_files = st.file_uploader("Upload CSV files", type=["csv"], accept_multiple_files=True)
    
    if uploaded_files:
        combined_df = pd.DataFrame()
        for uploaded_file in uploaded_files:
            df = pd.read_csv(uploaded_file, parse_dates=['Test_date_time'])
            df = clean_data(df)
            df["Source File"] = uploaded_file.name
            combined_df = pd.concat([combined_df, df], ignore_index=True)

        st.sidebar.header("Filters")
        start_date = st.sidebar.date_input("Start Date", combined_df["Test_date_time"].min())
        end_date = st.sidebar.date_input("End Date", (combined_df["Test_date_time"].max() + dt.timedelta(days=1)).date())
        # selected_parameters = st.sidebar.multiselect("Select Parameters",combined_df["Profile_id"].dropna().unique())
        
        filtered_df = combined_df[(combined_df['Test_date_time'] >= pd.to_datetime(start_date)) &
                                  (combined_df['Test_date_time'] <= pd.to_datetime(end_date))]
        # for parameter
        filtered_df["Profile_id"] = filtered_df["Profile_id"].astype(str)
        all_parameters = sorted(filtered_df["Profile_id"].dropna().unique().tolist())
        all_parameters.insert(0, "All parameters")
        selected_parameters = st.sidebar.multiselect("Select parameters", all_parameters, default=["All parameters"])
        if "All parameters" in selected_parameters:
            selected_parameters = all_parameters[1:]  # Exclude "All lots" from the selection
        filtered_df = filtered_df[filtered_df["Profile_id"].isin(selected_parameters)]

        # for lots
        # Ensure 'Lot' column has only string values to avoid mismatches
        filtered_df["Lot"] = filtered_df["Lot"].astype(str)
        all_lots = sorted(filtered_df["Lot"].dropna().unique().tolist())
        all_lots.insert(0, "All lots")
        selected_lots = st.sidebar.multiselect("Select Lots", all_lots, default=["All lots"])
        if "All lots" in selected_lots:
            selected_lots = all_lots[1:]  # Exclude "All lots" from the selection
        filtered_df = filtered_df[filtered_df["Lot"].isin(selected_lots)]

        # for serial no
        # Ensure 'Chip_serial_no' column is treated as a string to avoid mismatches
        filtered_df["Chip_serial_no"] = filtered_df["Chip_serial_no"].astype(str)
        all_series = sorted(filtered_df["Chip_serial_no"].dropna().unique().tolist())
        all_series.insert(0, "All series")
        selected_series = st.sidebar.multiselect("Select Series", all_series, default=["All series"])
        if "All series" in selected_series:
            selected_series = all_series[1:]  # Exclude "All series" from the selection
        filtered_df = filtered_df[filtered_df["Chip_serial_no"].isin(selected_series)]



        all_threshold = st.sidebar.number_input("Minimum 'All' Value", min_value=0, value=0)
    


        metric = "IND_per" if any(param in ["MTR", "INH"] for param in selected_parameters) else "INV_per"
        status_col = "Indeterminate" if metric == "IND_per" else "Invalid"
        
        index_columns = {
            "Lot Performance": ['Lot'],
            "Chip series": ['Chip_serial_no'],
            "Lot Chip series": ['Lot', 'Chip_serial_no'],
            "Lot chip batch chip series": ['Lot', 'Chip_batchno', 'Chip_serial_no'],
            "Detailed Data": ['Zone','State','Account Owner','Customer Type','Lab_name', 'Truelab_id', 'Lot', 'Chip_batchno', 'Chip_serial_no']
        }
        
        dataframes = {}
        tabs = st.tabs(list(index_columns.keys()))
        
        for tab, key in zip(tabs, index_columns.keys()):
            with tab:
                pivot = filtered_df.pivot_table(index=index_columns[key], values='Patient_id', columns='Test_status', aggfunc='count', margins=True)
                pivotdf = pd.DataFrame(pivot.to_records()).fillna(0)
                if status_col in pivotdf.columns and "All" in pivotdf.columns:
                    pivotdf[metric] = round(pivotdf[status_col] / pivotdf["All"] * 100, 2)
                dataframes[key] = pivotdf
                pivotdf = pivotdf[pivotdf['All'] > all_threshold]
                if "All" in pivotdf.columns:
                    pivotdf = pivotdf[pivotdf["All"] > all_threshold]  # Apply threshold filter only if column exists
                else:
                    st.warning("Column 'All' not found in pivot table. Check data consistency.")

                st.dataframe(pivotdf)
                pivotdf = pivotdf[pivotdf[index_columns[key][0]] != 'All']  # Remove rows where index is 'All'        
        # Create tabs
        inv_tab, ct_tab, time_tab = st.tabs(["INV_per/IND_per", "Ct values", "Time"])

        with inv_tab:
            # Define the columns to visualize
            columns_to_view = ['Zone', 'State', 'Account Owner', 'Customer Type', 'Lab_name', 
                                'Truelab_id', 'Lot', 'Chip_batchno', 'Chip_serial_no']

            # Check which column exists: "INV_per" or "IND_per"
            inv_col = "INV_per" if "INV_per" in pivotdf.columns else "IND_per" if "IND_per" in pivotdf.columns else None

            if inv_col:  # Proceed only if a valid column exists
                for column in columns_to_view:
                    if column in pivotdf.columns:
                        avg_inv_per = pivotdf.groupby(column)[inv_col].mean().reset_index()
                        fig = px.bar(avg_inv_per, x=column, y=inv_col, title=f"AVG {inv_col} Analysis - {column}")
                        st.plotly_chart(fig, use_container_width=True)
            else:
                st.warning("Neither 'INV_per' nor 'IND_per' exists in the dataset.")

        # Content for "Ct values" tab
        with ct_tab:
            # Convert Ct1, Ct2, Ct3 to numeric and replace non-numeric values with 0
            for i, col in zip(range(5, 9), ["Ct1", "Ct2", "Ct3"]):
                filtered_df[col] = pd.to_numeric(filtered_df[col], errors="coerce").fillna(0)

            # Create separate bell curves for Ct1, Ct2, and Ct3
            ct_columns = ["Ct1", "Ct2", "Ct3"]
            for ct_col in ct_columns:
                ct_values = filtered_df[filtered_df[ct_col] >= 10][ct_col]  # Filter values starting from 10

                # Create a histogram with KDE line
                fig = px.histogram(ct_values, x=ct_values, nbins=50, marginal="violin", opacity=0.7)
                
                # Update layout
                fig.update_layout(title=f"Bell Curve Distribution of {ct_col}",
                                xaxis_title=f"{ct_col} Values",
                                yaxis_title="Frequency",
                                template="plotly_white")

                # Show plot in Streamlit with a unique key
                st.plotly_chart(fig, key=f'key_{ct_col}_{i}')  # Ensure unique key for each chart

        # ---- Time TAB ---- #
        with time_tab:
            if "Test_date_time" in filtered_df.columns and "Test_status" in filtered_df.columns:
                # Convert 'Test_date_time' to datetime format
                filtered_df["Test_date_time"] = pd.to_datetime(filtered_df["Test_date_time"], errors="coerce")

                # Group data by time periods
                time_period = st.selectbox("Select Time Period", ["Daily", "Weekly", "Monthly", "Yearly"])

                # Define the time grouping
                if time_period == "Daily":
                    filtered_df["Period"] = filtered_df["Test_date_time"].dt.date
                elif time_period == "Weekly":
                    filtered_df["Period"] = filtered_df["Test_date_time"].dt.to_period("W").astype(str)
                elif time_period == "Monthly":
                    filtered_df["Period"] = filtered_df["Test_date_time"].dt.to_period("M").astype(str)
                else:  # Yearly
                    filtered_df["Period"] = filtered_df["Test_date_time"].dt.to_period("Y").astype(str)

                # Calculate total test count
                test_counts = filtered_df.groupby("Period")["Test_status"].count().reset_index(name="Total_Tests")

                # Determine if "Invalid" or "Indeterminate" exists
                status_to_check = None
                if "Invalid" in filtered_df["Test_status"].unique():
                    status_to_check = "Invalid"
                elif "Indeterminate" in filtered_df["Test_status"].unique():
                    status_to_check = "Indeterminate"

                if status_to_check:
                    # Calculate specific status count
                    status_counts = filtered_df[filtered_df["Test_status"] == status_to_check] \
                        .groupby("Period")["Test_status"].count().reset_index(name=f"{status_to_check}_Tests")

                    # Merge with total tests
                    time_grouped = test_counts.merge(status_counts, on="Period", how="left").fillna(0)

                    # Calculate percentage
                    time_grouped[f"{status_to_check}_Percentage"] = (time_grouped[f"{status_to_check}_Tests"] / time_grouped["Total_Tests"]) * 100

                    # Create line chart
                    fig = px.line(time_grouped, x="Period", y=f"{status_to_check}_Percentage", markers=True, 
                                title=f"{status_to_check} Test Percentage Over Time ({time_period})",
                                labels={"Period": time_period, f"{status_to_check}_Percentage": f"{status_to_check} Test Percentage (%)"})

                    # Show plot
                    st.plotly_chart(fig, use_container_width=True)


        excel_file = generate_excel(dataframes)
        st.sidebar.download_button("📥 Download Excel File", data=excel_file, file_name="data_export.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    main()
